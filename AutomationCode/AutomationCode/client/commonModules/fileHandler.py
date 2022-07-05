import json
from PyQt5 import QtWidgets
from glob import glob
import traceback
import pandas as pd
import openpyxl
import io



class fileHandler:
    def __init__(self):
        self.workBook = None
        self.sheetObj = None

    def browseExcelFile(self):
        try:
            selected, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Select Excel Files", '',
                                                                "Excel Files(*.xlsx)", )
            if selected:
                return selected
            else:
                return None
        except:
            print(traceback.format_exc())
            return None

    def readExcelFile(self, filePath):
        try:
            # # data = pd.read_excel(filePath, engine='openpyxl')
            # # jsonData = json.loads(data.to_json(orient='records'))
            # # return jsonData
            # workBook = openpyxl.load_workbook(filename=filePath, read_only=True)
            # sheetObj = workBook.active
            # maxCol = sheetObj.max_column + 1
            # maxRow = sheetObj.max_row + 1
            # jsonData = []
            # keys = []
            # for col in range(1, maxCol):
            #     value = sheetObj.cell(row=1, column=col).value
            #     keys.append(value)
            # for row in range(2, maxRow):
            #     temp = {}
            #     firstIndex = sheetObj.cell(row=row, column=1).value
            #     if firstIndex != None:
            #         for col in range(1, maxCol):
            #             value = sheetObj.cell(row=row, column=col).value
            #             temp.update({keys[col - 1]: value})
            #         jsonData.append(temp)
            # workBook.close()
            data = pd.read_excel(filePath, engine='openpyxl', dtype="str")
            jsonData = json.loads(data.to_json(orient='records'))
            return jsonData
        except:
            print(traceback.format_exc())
            return []

    def readExcelFileFromPandas(self, filePath):
        try:
            data = pd.read_excel(filePath, engine='openpyxl', dtype="str")
            jsonData = json.loads(data.to_json(orient='records'))
            return jsonData

        except:
            print(traceback.format_exc())
            return []

    def readExcelContent(self, filePath):
        try:
            with open(filePath, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            self.workBook = openpyxl.load_workbook(filename=in_mem_file, read_only=True)
            self.sheetObj = self.workBook.active
            maxCol = self.sheetObj.max_column + 1
            maxRow = self.sheetObj.max_row + 1
            keys = []
            for col in range(1, maxCol):
                value = self.sheetObj.cell(row=1, column=col).value
                keys.append(value)
            return keys, maxRow, maxCol
        except:
            print(traceback.format_exc())
            return [], []

    def getRowContent(self, rowNumber, maxCol, keys):
        try:
            temp = {}
            for col in range(1, maxCol):
                value = self.sheetObj.cell(row=rowNumber, column=col).value
                if value == None and col == 1:
                    break
                temp.update({keys[col - 1]: value})
            return temp
        except:
            print(traceback.format_exc())
            return {}

    def browseFile(self, method):
        try:
            selected = None
            if method == "file":
                selected, _ = QtWidgets.QFileDialog.getOpenFileNames(None, "Select JSON Files", '',
                                                                     "JSON Files(*.json)", )
            if method == "folder":
                selected = QtWidgets.QFileDialog.getExistingDirectory()
            if selected:
                if method == "file":
                    return selected
                elif method == "folder":
                    files = [f for f in glob(selected + "**/**/*.json", recursive=True)]
                    return files
                else:
                    return []
            else:
                return []
        except:
            print(traceback.format_exc())
            return []

    def readJson(self, filePath):
        try:
            # print(filePath)
            path = str(filePath).split("/")[-1]
            path = path.split(".")[0]
            with open(filePath, 'r', encoding='utf-8') as f:
                return path, json.loads(f.read())
        except:
            print(traceback.format_exc())
            return "", {}

    def generateExcelFile(self, content, fileName):
        try:
            df = pd.DataFrame(content)
            df.to_excel(fileName, index=False)

        except:
            print(traceback.format_exc())

    def generateJsonFile(self, content, fileName):
        try:
            with open(fileName, "a", encoding="utf-8") as f:
                f.write(json.dumps(content, indent=4))
        except:
            print(traceback.format_exc())

# fr = fileHandler()
# # maxRow = fr.readExcelFileFromPandas("E:/AdditLabs_Mahesh/Aquila Test Automation/client/inputFiles/Commands_Master.xlsx")
# # # print( maxRow)
# # for data in maxRow:
# #     print(data)
# # # print(fr.workBook, fr.sheetObj)
#
# fr.generateJsonFile({"Dependency":None,"TestCases":{"InputKey":"103","InputValue":"","ExpectedOutput":1,"DefaultValue":"taisysnet","RangeValue":"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa"}}, "E:/AdditLabs_Mahesh/Aquila Test Automation/client/outputFiles/commadScripts/BC/147/147_0_123467.json")
